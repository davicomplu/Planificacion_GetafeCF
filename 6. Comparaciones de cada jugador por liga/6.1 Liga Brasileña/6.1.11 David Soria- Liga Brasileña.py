#David Soria- Liga Brasileña (creación de CSVs y generación de archivos excel)


import pandas as pd
import os

# Paso 1: Leer el archivo CSV inicial
ruta_inicial = r"C:/Users/david/LIGAS/Getafe 18-19/Escalado_Getafe 18-19/Jugadores/PO. David Soria/David Soria.csv"
df_inicial = pd.read_csv(ruta_inicial)

# Limpiar datos del archivo inicial
columnas_relevantes = ["Jugador", "Mín",'GC90', '% Salvadas', 'PaC%', '% Salvadas.1','PSxG+/-', 'Long. prom..1', '% Cmp.1', '% Cmp.2', '% Cmp.3', 'Err']
df_inicial = df_inicial[columnas_relevantes]

# Consolidar datos numéricos en una sola fila
fila_consolidada = {}
for col in columnas_relevantes:
    if col == "Jugador":
        fila_consolidada[col] = "David Soria"
    else:
        fila_consolidada[col] = df_inicial[col].dropna().iloc[0] if not df_inicial[col].dropna().empty else ""

df_inicial_limpio = pd.DataFrame([fila_consolidada])

# Paso 2: Procesar cada equipo
equipos = [
    "Athletico-Paranaense", "Atletico-Goianiense", "Atletico-Mineiro", "Bahia", "Botafogo", "Red-Bull-Bragantino", "Corinthians", "Criciuma", "Cruzeiro",
    "Cuiaba", "Flamengo", "Fluminense", "Fortaleza", "Gremio", "Internacional", "Juventude", "Palmeiras", "Sao-Paulo", "Vasco-da-Gama", "Vitoria"
]

excluir_archivos = ["Série A, Série A.csv", "Marcadores y partidos.csv"]

for equipo in equipos:
    ruta_equipo = rf"C:/Users/david/LIGAS/Brasileirao/{equipo}"
    
    # Leer todos los archivos CSV excepto los excluidos
    dataframes = []
    for archivo in os.listdir(ruta_equipo):
        if archivo.endswith(".csv") and archivo not in excluir_archivos:
            df_temp = pd.read_csv(os.path.join(ruta_equipo, archivo))
            dataframes.append(df_temp)

    # Concatenar todos los DataFrames
    df_equipo = pd.concat(dataframes, ignore_index=True)

    # Filtrar filas donde la columna "Posc" contiene "DF"
    df_filtrado = df_equipo[df_equipo["Posc"].str.contains("PO", na=False)]

    # Seleccionar solo las columnas relevantes
    df_filtrado = df_filtrado[columnas_relevantes]

    # Solucionar duplicados en la columna "Jugador" y alinear datos en una sola fila por jugador
    df_filtrado_grouped = df_filtrado.groupby("Jugador").first().reset_index()

    # Generar el archivo CSV final
    ruta_salida = r"C:/Users/david/LIGAS/Comparación/Brasileirao/PO. David Soria/"
    os.makedirs(ruta_salida, exist_ok=True)
    archivo_salida = os.path.join(ruta_salida, f"PO. David Soria-{equipo}.csv")

    # Agregar datos del archivo inicial al nuevo CSV
    df_final = pd.concat([df_inicial_limpio, df_filtrado_grouped], ignore_index=True)

    # Subir datos numéricos a la fila correspondiente y exportar a CSV
    df_final.to_csv(archivo_salida, index=False)

print("Proceso completado. Se han generado archivos CSV para todos los equipos.")

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Directorios (verifica que las rutas sean correctas)
directorio_csv = r"C:/Users/david/LIGAS/Comparación/Brasileirao/PO. David Soria/"
directorio_excel = r"C:/Users/david/OneDrive/Escritorio/MÁSTER/TFM/RESULTADOS COMPARACIÓN/PO. David Soria/Brasileirao/"

# Asegurar que el directorio existe (esto es clave)
os.makedirs(directorio_excel, exist_ok=True)

# Procesar archivos
for archivo_csv in os.listdir(directorio_csv):
    if archivo_csv.endswith(".csv"):
        # 1. Leer CSV
        ruta_csv = os.path.join(directorio_csv, archivo_csv)
        try:
            df = pd.read_csv(ruta_csv)
        except Exception as e:
            print(f"Error leyendo {archivo_csv}: {e}")
            continue

        # 2. Crear nombre para el Excel
        nombre_equipo = archivo_csv.split("-")[-1].replace(".csv", "")
        archivo_excel = f"PO. David Soria-{nombre_equipo}.xlsx"
        ruta_excel = os.path.join(directorio_excel, archivo_excel)

        # 3. Convertir a Excel
        try:
            df.to_excel(ruta_excel, index=False, engine='openpyxl')
            
            # 4. Aplicar formato
            wb = load_workbook(ruta_excel)
            ws = wb.active
            
            # Colores
            verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Headers
            headers_relevantes = ['GC90', '% Salvadas', 'PaC%', '% Salvadas.1','PSxG+/-', 
                                'Long. prom..1', '% Cmp.1', '% Cmp.2', '% Cmp.3', 'Err']
            
            # Encontrar columnas
            col_indices = {}
            col_jugador = None
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header in headers_relevantes:
                    col_indices[header] = col
                elif header == "Jugador":
                    col_jugador = col
            
            if not col_jugador:
                print(f"Advertencia: No se encontró columna 'Jugador' en {archivo_csv}")
                continue

            # Procesar datos
            for row in range(3, ws.max_row + 1):
                total_verde = 0
                for header, col in col_indices.items():
                    referencia = ws.cell(row=2, column=col).value
                    valor = ws.cell(row=row, column=col).value
                    
                    if isinstance(referencia, (int, float)) and isinstance(valor, (int, float)):
                        if header in ["Err", "GC90"]:  # Menor es mejor
                            if valor <= referencia:
                                ws.cell(row=row, column=col).fill = verde
                                total_verde += 1
                            else:
                                ws.cell(row=row, column=col).fill = rojo
                        else:  # Mayor es mejor
                            if valor >= referencia:
                                ws.cell(row=row, column=col).fill = verde
                                total_verde += 1
                            else:
                                ws.cell(row=row, column=col).fill = rojo
                
                # Colorear jugador
                if total_verde >= len(headers_relevantes) * (2 / 3):
                    ws.cell(row=row, column=col_jugador).fill = verde
                else:
                    ws.cell(row=row, column=col_jugador).fill = rojo
            
            wb.save(ruta_excel)
            print(f"Archivo creado: {ruta_excel}")
            
        except Exception as e:
            print(f"Error procesando {archivo_csv}: {e}")

print("Proceso finalizado. Verifica los archivos en:", directorio_excel)