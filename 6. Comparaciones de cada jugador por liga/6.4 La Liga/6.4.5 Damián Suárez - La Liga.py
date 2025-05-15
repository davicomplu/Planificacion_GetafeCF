#Damián Suárez - La Liga (creación de CSVs y generación de archivos excel)


import pandas as pd
import os

# Paso 1: Leer el archivo CSV inicial
ruta_inicial = r"C:/Users/david/LIGAS/Getafe 18-19/Escalado_Getafe 18-19/Jugadores/LD. Damián Suárez/damian_suarez.csv"
df_inicial = pd.read_csv(ruta_inicial)

# Limpiar datos del archivo inicial
# Mantener únicamente las columnas relevantes
columnas_relevantes = ["Jugador", "Mín", "TA", "TR", "G-TP.1", "Ast", "SCA90", "GCA90", "Tkl%", "Tkl+Int", "Err", "% de ganados"]
df_inicial = df_inicial[columnas_relevantes]

# Consolidar datos numéricos en una sola fila
# Para cada columna relevante, tomar el primer valor no vacío
fila_consolidada = {}
for col in columnas_relevantes:
    if col == "Jugador":
        fila_consolidada[col] = "Damián Suárez"  # Mantener solo el nombre del jugador
    else:
        fila_consolidada[col] = df_inicial[col].dropna().iloc[0] if not df_inicial[col].dropna().empty else ""

# Crear un DataFrame con la fila consolidada
df_inicial_limpio = pd.DataFrame([fila_consolidada])

# Paso 2: Crear un DataFrame con los archivos CSV de la segunda ruta
ruta_real_madrid = r"C:/Users/david/LIGAS/La Liga/Alavés/Escalado_Alavés/"
excluir_archivos = ["La Liga.csv", "Marcadores y partidos.csv", "Portería avanzada.csv", "Porteros.csv"]

# Leer todos los archivos CSV excepto los excluidos
dataframes = []
for archivo in os.listdir(ruta_real_madrid):
    if archivo.endswith(".csv") and archivo not in excluir_archivos:
        df_temp = pd.read_csv(os.path.join(ruta_real_madrid, archivo))
        dataframes.append(df_temp)

# Concatenar todos los DataFrames
df_real_madrid = pd.concat(dataframes, ignore_index=True)

# Paso 3: Filtrar filas donde la columna "Posc" contiene "DF"
df_filtrado = df_real_madrid[df_real_madrid["Posc"].str.contains("DF", na=False)]

# Seleccionar solo las columnas relevantes
df_filtrado = df_filtrado[columnas_relevantes]

# Solucionar duplicados en la columna "Jugador" y alinear datos en una sola fila por jugador
df_filtrado_grouped = df_filtrado.groupby("Jugador").first().reset_index()

# Paso 4: Generar el archivo CSV final
ruta_salida = r"C:/Users/david/LIGAS/Comparación/La Liga/LD. Damián Suárez/"
os.makedirs(ruta_salida, exist_ok=True)
archivo_salida = os.path.join(ruta_salida, "LD. Damián Suárez-Alavés.csv")

# Agregar datos del archivo inicial al nuevo CSV
df_final = pd.concat([df_inicial_limpio, df_filtrado_grouped], ignore_index=True)

# Subir datos numéricos a la fila correspondiente y exportar a CSV
df_final.to_csv(archivo_salida, index=False)


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paso 1: Leer el archivo CSV generado previamente
ruta_csv = r"C:/Users/david/LIGAS/Comparación/La Liga/LD. Damián Suárez/LD. Damián Suárez-Alavés.csv"
df = pd.read_csv(ruta_csv)

# Paso 2: Convertir el archivo CSV a Excel
ruta_excel = r"C:/Users/david/OneDrive/Escritorio/MÁSTER/TFM/RESULTADOS COMPARACIÓN/LD. Damián Suárez/LD. Damián Suárez-Alavés.xlsx"
df.to_excel(ruta_excel, index=False, engine='openpyxl')

# Paso 3: Aplicar colores según las reglas
# Cargar el archivo Excel
wb = load_workbook(ruta_excel)
ws = wb.active

# Definir colores
verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Headers relevantes para la comparación
headers_relevantes = ["G-TP.1", "Ast", "SCA90", "GCA90", "Tkl%", "Tkl+Int", "Err", "% de ganados"]

# Obtener índices de las columnas relevantes y de la columna "Jugador"
col_indices = {}
col_jugador = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value  # Obtener el valor del header
    if header in headers_relevantes:
        col_indices[header] = col
    if header == "Jugador":
        col_jugador = col

# Pintar la fila 2 de verde (referencia)
for col in col_indices.values():
    ws.cell(row=2, column=col).fill = verde

# Comparar datos de las filas restantes con los datos de la fila 2
for row in range(3, ws.max_row + 1):  # Comenzamos desde la fila 3
    total_verde = 0  # Contador para celdas verdes en esta fila
    for header, col in col_indices.items():
        referencia = ws.cell(row=2, column=col).value  # Valor de referencia (fila 2)
        valor_actual = ws.cell(row=row, column=col).value  # Valor actual

        # Asegurarse de que ambos valores sean numéricos antes de comparar
        if isinstance(referencia, (int, float)) and isinstance(valor_actual, (int, float)):
            if header == "Err":  # Revertir lógica para la columna "Err"
                if valor_actual >= referencia:
                    ws.cell(row=row, column=col).fill = rojo
                else:
                    ws.cell(row=row, column=col).fill = verde
            else:  # Lógica normal para las demás columnas
                if valor_actual >= referencia:
                    ws.cell(row=row, column=col).fill = verde
                    total_verde += 1  # Incrementar contador si la celda es verde
                else:
                    ws.cell(row=row, column=col).fill = rojo

    # Colorear la celda en la columna "Jugador" según la condición de 2/3 verdes
    if total_verde >= len(headers_relevantes) * (2 / 3):  # Al menos 2/3 verdes (excepto "Err")
        ws.cell(row=row, column=col_jugador).fill = verde
    else:
        ws.cell(row=row, column=col_jugador).fill = rojo

# Guardar cambios en el archivo Excel
wb.save(ruta_excel)
