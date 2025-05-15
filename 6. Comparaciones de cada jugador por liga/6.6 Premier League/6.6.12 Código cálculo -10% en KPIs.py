#Código utilizado para calcular aquellos jugadores cuyos datos están solo un 10% por debajo de los datos del jugador modelo (de un jugador a otro únicamente debemos cambiar el nombre y los KPIs seleccionados)


import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Directorios (actualizados según lo indicado)
directorio_csv = r"C:/Users/david/LIGAS/Comparación/PL/DC. Jaime Mata/"
directorio_excel = r"C:/Users/david/OneDrive/Escritorio/MÁSTER/TFM/RESULTADOS COMPARACIÓN/DC. Jaime Mata/PL/10%/"

# Asegurar que el directorio de salida existe
os.makedirs(directorio_excel, exist_ok=True)

# Definir colores
verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Headers relevantes para la comparación
headers_relevantes = [
    '% de ganados', 'G-TP.1', 'Ast', 'SCA90', 'GCA90', '% Cmp.1', 
    '% Cmp.2', '% Cmp.3', 'Exitosa%', 'TalArc/90', 'G/TalArc', 
    'npxG', 'Tkl+Int', 'Tkl%'
]

# Procesar cada archivo CSV
for archivo_csv in os.listdir(directorio_csv):
    if archivo_csv.endswith(".csv"):
        # Leer CSV y generar nombre de archivo Excel
        ruta_csv = os.path.join(directorio_csv, archivo_csv)
        nombre_equipo = archivo_csv.split("-")[-1].replace(".csv", "")
        archivo_excel = f"DC. Jaime Mata-{nombre_equipo}.xlsx"
        ruta_excel = os.path.join(directorio_excel, archivo_excel)
        
        # Convertir CSV a Excel
        df = pd.read_csv(ruta_csv)
        df.to_excel(ruta_excel, index=False, engine='openpyxl')
        
        # Aplicar estilos
        wb = load_workbook(ruta_excel)
        ws = wb.active
        
        # Obtener índices de columnas
        col_indices = {}
        col_jugador = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in headers_relevantes:
                col_indices[header] = col
            elif header == "Jugador":
                col_jugador = col
        
        # Pintar fila de referencia (fila 2) en verde
        for col in col_indices.values():
            ws.cell(row=2, column=col).fill = verde
        
        # Evaluar cada jugador (filas 3 en adelante)
        for row in range(3, ws.max_row + 1):
            total_verdes = 0
            total_amarillas = 0
            
            for header, col in col_indices.items():
                referencia = ws.cell(row=2, column=col).value
                valor = ws.cell(row=row, column=col).value
                
                if isinstance(referencia, (int, float)) and isinstance(valor, (int, float)):
                    umbral_90 = referencia * 0.9  # 90% del valor de referencia
                    
                    # Coloración de celdas individuales
                    if header == "Err":  # Caso especial (menor es mejor)
                        if valor <= referencia:
                            ws.cell(row=row, column=col).fill = verde
                            total_verdes += 1
                        elif valor <= referencia * 1.1:  # +10% para errores
                            ws.cell(row=row, column=col).fill = amarillo
                            total_amarillas += 1
                        else:
                            ws.cell(row=row, column=col).fill = rojo
                    else:  # Para el resto (mayor es mejor)
                        if valor >= referencia:
                            ws.cell(row=row, column=col).fill = verde
                            total_verdes += 1
                        elif valor >= umbral_90:
                            ws.cell(row=row, column=col).fill = amarillo
                            total_amarillas += 1
                        else:
                            ws.cell(row=row, column=col).fill = rojo
            
            # Coloración del nombre del jugador (prioridad: verde > amarillo > rojo)
            total_metricas = len(headers_relevantes)
            if total_verdes / total_metricas >= 2/3:
                ws.cell(row=row, column=col_jugador).fill = verde
            elif (total_verdes + total_amarillas) / total_metricas >= 2/3 and total_amarillas >= 1:
                ws.cell(row=row, column=col_jugador).fill = amarillo
            else:
                ws.cell(row=row, column=col_jugador).fill = rojo
        
        # Guardar cambios
        wb.save(ruta_excel)

print("Proceso completado. Archivos Excel generados en:", directorio_excel)
