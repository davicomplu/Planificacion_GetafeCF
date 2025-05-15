#Arambarri - Premier League (creación de CSVs y generación de archivos excel)


import pandas as pd
import os

# Paso 1: Leer el archivo CSV inicial
ruta_inicial = r"C:/Users/david/LIGAS/Getafe 18-19/Escalado_Getafe 18-19/Jugadores/MC. Mauro Arambarri/Mauro Arambarri.csv"
df_inicial = pd.read_csv(ruta_inicial)

# Limpiar datos del archivo inicial
columnas_relevantes = ["Jugador", "Mín", "TA", "TR", "G-TP.1", "Ast",'% Cmp.1', '% Cmp.2', '% Cmp.3', "SCA90", "GCA90", "Tkl%", "Tkl+Int", "Err", "% de ganados"]
df_inicial = df_inicial[columnas_relevantes]

# Consolidar datos numéricos en una sola fila
fila_consolidada = {}
for col in columnas_relevantes:
    if col == "Jugador":
        fila_consolidada[col] = "Mauro Arambarri"
    else:
        fila_consolidada[col] = df_inicial[col].dropna().iloc[0] if not df_inicial[col].dropna().empty else ""

df_inicial_limpio = pd.DataFrame([fila_consolidada])

# Paso 2: Procesar cada equipo
equipos = [
    "Arsenal", "Aston Villa", "Bournemouth", "Brentford", "Brighton", 
    "Chelsea", "Crystal Palace", "Everton", "Fulham", "Ipswich Town", 
    "Leicester City", "Liverpool", "Manchester City", "Manchester Utd", 
    "Newcastle Utd", "Nottingham Forest", "Southampton", "Tottenham", 
    "West Ham", "Wolves"
]

excluir_archivos = ["Premier League.csv", "Marcadores y partidos.csv", "Portería avanzada.csv", "Porteros.csv"]

for equipo in equipos:
    ruta_equipo = rf"C:/Users/david/LIGAS/PL/{equipo}/Escalado_{equipo}/"
    
    # Leer todos los archivos CSV excepto los excluidos
    dataframes = []
    for archivo in os.listdir(ruta_equipo):
        if archivo.endswith(".csv") and archivo not in excluir_archivos:
            df_temp = pd.read_csv(os.path.join(ruta_equipo, archivo))
            dataframes.append(df_temp)

    # Concatenar todos los DataFrames
    df_equipo = pd.concat(dataframes, ignore_index=True)

    # Filtrar filas donde la columna "Posc" contiene "DF"
    df_filtrado = df_equipo[df_equipo["Posc"].str.contains("CC", na=False)]

    # Seleccionar solo las columnas relevantes
    df_filtrado = df_filtrado[columnas_relevantes]

    # Solucionar duplicados en la columna "Jugador" y alinear datos en una sola fila por jugador
    df_filtrado_grouped = df_filtrado.groupby("Jugador").first().reset_index()

    # Generar el archivo CSV final
    ruta_salida = r"C:/Users/david/LIGAS/Comparación/PL/MC. Mauro Arambarri/"
    os.makedirs(ruta_salida, exist_ok=True)
    archivo_salida = os.path.join(ruta_salida, f"MC. Mauro Arambarri-{equipo}.csv")

    # Agregar datos del archivo inicial al nuevo CSV
    df_final = pd.concat([df_inicial_limpio, df_filtrado_grouped], ignore_index=True)

    # Subir datos numéricos a la fila correspondiente y exportar a CSV
    df_final.to_csv(archivo_salida, index=False)

print("Proceso completado. Se han generado archivos CSV para todos los equipos.")

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Directorio donde se encuentran los archivos CSV
directorio_csv = r"C:/Users/david/LIGAS/Comparación/PL/MC. Mauro Arambarri/"

# Directorio donde se guardarán los archivos Excel
directorio_excel = r"C:/Users/david/OneDrive/Escritorio/MÁSTER/TFM/RESULTADOS COMPARACIÓN/MC. Mauro Arambarri/PL/"

# Asegurarse de que el directorio de salida existe
os.makedirs(directorio_excel, exist_ok=True)

# Procesar cada archivo CSV en el directorio
for archivo_csv in os.listdir(directorio_csv):
    if archivo_csv.endswith(".csv"):
        # Paso 1: Leer el archivo CSV
        ruta_csv = os.path.join(directorio_csv, archivo_csv)
        df = pd.read_csv(ruta_csv)

        # Obtener el nombre del equipo
        nombre_equipo = archivo_csv.split("-")[-1].replace(".csv", "")

        # Paso 2: Convertir el archivo CSV a Excel
        archivo_excel = f"MC. Mauro Arambarri-{nombre_equipo}.xlsx"
        ruta_excel = os.path.join(directorio_excel, archivo_excel)
        df.to_excel(ruta_excel, index=False, engine='openpyxl')

        # Paso 3: Aplicar colores según las reglas
        wb = load_workbook(ruta_excel)
        ws = wb.active

        # Definir colores
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Headers relevantes para la comparación
        headers_relevantes = ["G-TP.1", "Ast", '% Cmp.1', '% Cmp.2', '% Cmp.3', "SCA90", "GCA90", "Tkl%", "Tkl+Int", "Err", "% de ganados"]

        # Obtener índices de las columnas relevantes y de la columna "Jugador"
        col_indices = {}
        col_jugador = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in headers_relevantes:
                col_indices[header] = col
            if header == "Jugador":
                col_jugador = col

        # Pintar la fila 2 de verde (referencia)
        for col in col_indices.values():
            ws.cell(row=2, column=col).fill = verde

        # Comparar datos de las filas restantes con los datos de la fila 2
        for row in range(3, ws.max_row + 1):
            total_verde = 0
            for header, col in col_indices.items():
                referencia = ws.cell(row=2, column=col).value
                valor_actual = ws.cell(row=row, column=col).value

                if isinstance(referencia, (int, float)) and isinstance(valor_actual, (int, float)):
                    if header == "Err":
                        if valor_actual <= referencia:
                            ws.cell(row=row, column=col).fill = verde
                            total_verde += 1
                        else:
                            ws.cell(row=row, column=col).fill = rojo        
                    else:
                        if valor_actual >= referencia:
                            ws.cell(row=row, column=col).fill = verde
                            total_verde += 1
                        else:
                            ws.cell(row=row, column=col).fill = rojo

            # Colorear la celda en la columna "Jugador" según la condición de 2/3 verdes
            if total_verde >= len(headers_relevantes) * (2 / 3):
                ws.cell(row=row, column=col_jugador).fill = verde
            else:
                ws.cell(row=row, column=col_jugador).fill = rojo

        # Guardar cambios en el archivo Excel
        wb.save(ruta_excel)

print("Proceso completado. Se han generado archivos Excel para todos los equipos.")
